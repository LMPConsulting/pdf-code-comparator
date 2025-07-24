# src/reporting.py
import pandas as pd
import os
from datetime import datetime

def generate_report(in_both, only_in_pdf1, only_in_pdf2, output_dir, report_format):
    """
    Erstellt einen Report basierend auf den Code-Vergleichsergebnissen.

    Args:
        in_both (set): Codes, die in beiden PDFs gefunden und validiert wurden.
        only_in_pdf1 (set): Codes, die nur im ersten PDF gefunden und validiert wurden.
        only_in_pdf2 (set): Codes, die nur im zweiten PDF gefunden und validiert wurden.
        output_dir (str): Verzeichnis, in dem der Report gespeichert werden soll.
        report_format (str): 'xlsx' oder 'csv'.

    Returns:
        str: Der vollständige Pfad zur erstellten Reportdatei oder None im Fehlerfall.
    """
    print(f"Erstelle Report im Format: {report_format} im Ordner {output_dir}")

    # Sortieren der Codes für eine bessere Lesbarkeit
    codes_in_both = sorted(list(in_both))
    codes_only_in_pdf1 = sorted(list(only_in_pdf1))
    codes_only_in_pdf2 = sorted(list(only_in_pdf2))

    # Aktuellen Zeitstempel für den Dateinamen
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = f"Code_Comparison_Report_{timestamp}"

    # Erstelle Zielordner, falls nicht vorhanden (wichtig, falls das Ziel ein Unterordner ist)
    os.makedirs(output_dir, exist_ok=True)


    try:
        if report_format == 'xlsx':
            # Report als Excel-Datei
            filename = f"{base_filename}.xlsx"
            report_path = os.path.join(output_dir, filename)

            # DataFrames für jede Menge erstellen
            df_both = pd.DataFrame(codes_in_both, columns=["Codes in beiden Dokumenten"])
            df_only1 = pd.DataFrame(codes_only_in_pdf1, columns=["Codes nur in Dokument 1"])
            df_only2 = pd.DataFrame(codes_only_in_pdf2, columns=["Codes nur in Dokument 2"])

            # Nur Report erstellen, wenn es überhaupt Codes gab (siehe main.py Logik)
            # if df_both.empty and df_only1.empty and df_only2.empty:
            #      print("Keine Codes zum Reporten gefunden.")
            #      return None # Dies wird bereits in main.py abgefangen, sollte hier nicht passieren

            # Excel Writer erstellen
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # Schreiben Sie nur Sheets, wenn die entsprechenden Mengen nicht leer sind
                if not df_both.empty:
                    df_both.to_excel(writer, sheet_name='Codes in beiden', index=False)
                if not df_only1.empty:
                    df_only1.to_excel(writer, sheet_name='Codes nur in PDF 1', index=False)
                if not df_only2.empty:
                    df_only2.to_excel(writer, sheet_name='Codes nur in PDF 2', index=False)

            print(f"Excel Report erstellt: {report_path}")
            return report_path

        elif report_format == 'csv':
            # Report als CSV-Datei (einfacher, listet die Mengen untereinander)
            filename = f"{base_filename}.csv"
            report_path = os.path.join(output_dir, filename)

            # Für Einfachheit: Eine CSV, die die Mengen listet
            with open(report_path, 'w', newline='', encoding='utf-8') as f:
                # Schreiben Sie nur, wenn die entsprechenden Mengen nicht leer sind
                if codes_in_both:
                    f.write("Codes in beiden Dokumenten:\n")
                    for code in codes_in_both:
                        f.write(f"{code}\n")
                    f.write("\n") # Leerzeile zwischen Abschnitten

                if codes_only_in_pdf1:
                    f.write("Codes nur in Dokument 1:\n")
                    for code in codes_only_in_pdf1:
                        f.write(f"{code}\n")
                    f.write("\n")

                if codes_only_in_pdf2:
                    f.write("Codes nur in Dokument 2:\n")
                    for code in codes_only_in_pdf2:
                        f.write(f"{code}\n")

                # Wenn alle Listen leer waren, schreiben Sie eine kurze Info
                if not (codes_in_both or codes_only_in_pdf1 or codes_only_in_pdf2):
                     f.write("Keine Codes gefunden, die der Masterliste entsprechen.\n")


            print(f"CSV Report erstellt: {report_path}")
            return report_path

        else:
            print(f"FEHLER: Ungültiges Reportformat in Konfiguration: {report_format}. Unterstützt: xlsx, csv")
            return None

    except Exception as e:
        print(f"FEHLER beim Erstellen des Reports: {e}")
        return None

# Beispielaufruf (wird nicht ausgeführt, wenn importiert)
if __name__ == '__main__':
    # Dummy-Daten für Test
    # Diese Sets stellen Codes DAR, die bereits extrahiert, bereinigt und validiert wurden
    dummy_in_both = {'A1X', 'P0B'}
    dummy_only1 = {'C3Z'}
    dummy_only2 = {'D4W'}

    # Report erstellen (im aktuellen Verzeichnis)
    current_dir = os.path.dirname(os.path.abspath(__file__)) # src Ordner
    # Gehe einen Ordner hoch, um in den Projekt-Basisordner zu schreiben
    project_dir = os.path.abspath(os.path.join(current_dir, '..'))

    print("--- Test reporting.py Funktionen ---")
    report_path_xlsx = generate_report(dummy_in_both, dummy_only1, dummy_only2, project_dir, 'xlsx')
    report_path_csv = generate_report(dummy_in_both, dummy_only1, dummy_only2, project_dir, 'csv')
    report_path_empty = generate_report(set(), set(), set(), project_dir, 'xlsx') # Test mit leeren Mengen

    print(f"Dummy Report XLSX: {report_path_xlsx}")
    print(f"Dummy Report CSV: {report_path_csv}")
    print(f"Dummy Report Empty: {report_path_empty}")

def generate_enhanced_report(original_results, corrected_results, corrections, output_dir, report_format, raw_codes_pdf1=None, raw_codes_pdf2=None):
    """
    Generiert einen erweiterten Bericht mit bidirektionaler OCR-Korrektur und Kontext-Analyse.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if report_format.lower() == 'xlsx':
        filename = f"Code_Comparison_Report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Hauptergebnisse (optimiert sortiert und farbkodiert)
            main_results = []
            
            # Erstelle Mapping von Korrekturen (inklusive erweiterte OCR-Korrekturen)
            correction_map = {}
            for corr in corrections:
                # Alle Korrekturen einschließen (auch erweiterte OCR-Korrekturen)
                correction_map[corr['corrected']] = corr
                print(f"    Korrektur-Mapping: '{corr['corrected']}' -> {corr['correction_type']} ({corr['probability']:.0%})")
            
            print(f"  Gesamt Korrekturen im Mapping: {len(correction_map)}")
            print(f"  Correction-Map Keys: {list(correction_map.keys())}")
            
            # Erstelle Seiten-Mapping aus rohen Daten
            def create_page_mapping(raw_codes):
                page_map = {}
                original_page_map = {}  # Für ursprüngliche (unkorrigierte) Codes
                if raw_codes:
                    for code, page, pos in raw_codes:
                        from .core import clean_code
                        from .core import clean_code
                        cleaned = clean_code(code, None)  # Keine Master-Validierung im Reporting
                        
                        # Speichere bereinigten Code mit Seite
                        if cleaned not in page_map:
                            page_map[cleaned] = page
                        
                        # Speichere auch ursprünglichen (unkorrigierten) Code
                        if code.upper() not in original_page_map:
                            original_page_map[code.upper()] = page
                
                return page_map, original_page_map
            
            pdf1_page_map, pdf1_original_map = create_page_mapping(raw_codes_pdf1)
            pdf2_page_map, pdf2_original_map = create_page_mapping(raw_codes_pdf2)
            
            # Importiere is_control_code für Steuer-Code-Erkennung
            from .code_filters import is_control_code
            
            # 1. STEUER-CODES ZUERST (I-Codes) - alle Kategorien
            control_codes_both = corrected_results.get('control', {}).get('in_both', set())
            control_only_pdf1 = corrected_results.get('control', {}).get('only_in_pdf1', set())
            control_only_pdf2 = corrected_results.get('control', {}).get('only_in_pdf2', set())
            
            # Steuer-Codes in beiden PDFs (grau markiert)
            for code in control_codes_both:
                # Korrektur-Informationen
                if code in correction_map:
                    corr = correction_map[code]
                    korrektur_info = corr['method']
                    sicherheit = f"{corr['probability']:.0%}"
                    color_code = 'grau'  # Steuer-Codes immer grau
                    sort_priority = 0  # Ganz oben
                else:
                    # Prüfe ob es ein direkter Match ist
                    direct_match = next((c for c in corrections if c['corrected'] == code and c['correction_type'] == 'Direkter Match'), None)
                    if direct_match:
                        korrektur_info = direct_match['method']
                        sicherheit = f"{direct_match['probability']:.0%}"
                        color_code = 'grau'  # Steuer-Codes immer grau
                        sort_priority = 0  # Ganz oben
                    else:
                        korrektur_info = ''
                        sicherheit = 'N/A'
                        color_code = 'grau'  # Steuer-Codes immer grau
                        sort_priority = 0  # Ganz oben
                
                main_results.append({
                    'Code': code,
                    'Status': 'In beiden PDFs (Steuer-Code)',
                    'PDF1': 'Ja',
                    'PDF2': 'Ja',
                    'Sicherheit': sicherheit,
                    'Matching/Korrektur': korrektur_info,
                    'color_code': color_code,
                    'sort_priority': sort_priority,
                    'code_type': 'control'
                })
            
            # Steuer-Codes nur in PDF1 (grau markiert, mit 40% Sicherheit)
            for code in control_only_pdf1:
                main_results.append({
                    'Code': code,
                    'Status': 'Nur in PDF1 (Steuer-Code)',
                    'PDF1': 'Ja',
                    'PDF2': 'Nein',
                    'Sicherheit': '40%',  # 40% für Fund in Masterliste
                    'Matching/Korrektur': 'Steuer-Code nur in PDF1, in Masterliste validiert',
                    'color_code': 'grau',
                    'sort_priority': 0,  # Ganz oben
                    'code_type': 'control'
                })
            
            # Steuer-Codes nur in PDF2 (grau markiert, mit 40% Sicherheit)
            for code in control_only_pdf2:
                main_results.append({
                    'Code': code,
                    'Status': 'Nur in PDF2 (Steuer-Code)',
                    'PDF1': 'Nein',
                    'PDF2': 'Ja',
                    'Sicherheit': '40%',  # 40% für Fund in Masterliste
                    'Matching/Korrektur': 'Steuer-Code nur in PDF2, in Masterliste validiert',
                    'color_code': 'grau',
                    'sort_priority': 0,  # Ganz oben
                    'code_type': 'control'
                })
            
            # 2. DANN NORMALE CODES
            normal_codes_both = corrected_results.get('normal', {}).get('in_both', set())
            
            # Normale Codes in beiden PDFs
            for code in normal_codes_both:
                print(f"    Prüfe Code '{code}' in correction_map: {code in correction_map}")
                # Korrektur-Informationen (inklusive erweiterte OCR-Korrekturen)
                if code in correction_map:
                    corr = correction_map[code]
                    # Verwende nur die method (detaillierter Kommentar)
                    korrektur_info = corr['method']
                    sicherheit = f"{corr['probability']:.0%}"
                    # Seiten-Information wird nicht mehr benötigt
                    
                    # Aktualisierte Farbkodierung
                    if corr['probability'] >= 1.0:
                        color_code = 'dunkel_gruen'  # Dunkel Grün für 100%
                    elif corr['probability'] >= 0.90:
                        color_code = 'hell_gruen'  # Helles Grün für 99-90%
                    elif corr['probability'] >= 0.80:
                        color_code = 'dunkel_gelb'  # Dunkles Gelb für 89-80%
                    elif corr['probability'] >= 0.60:
                        color_code = 'hell_gelb'  # Helles Gelb für 79-60%
                    else:
                        color_code = 'orange'  # Orange für unter 60%
                else:
                    # Prüfe ob es ein direkter Match ist
                    direct_match = next((c for c in corrections if c['corrected'] == code and c['correction_type'] == 'Direkter Match'), None)
                    if direct_match:
                        # Begründung für die Sicherheit
                        korrektur_info = direct_match['method']
                        sicherheit = f"{direct_match['probability']:.0%}"
                        seite = ''
                        
                        # Aktualisierte Farbkodierung für direkte Matches
                        prob = direct_match['probability']
                        if prob >= 1.0:
                            color_code = 'dunkel_gruen'  # Dunkel Grün für 100%
                        elif prob >= 0.90:
                            color_code = 'hell_gruen'  # Helles Grün für 99-90%
                        elif prob >= 0.80:
                            color_code = 'dunkel_gelb'  # Dunkles Gelb für 89-80%
                        elif prob >= 0.60:
                            color_code = 'hell_gelb'  # Helles Gelb für 79-60%
                        else:
                            color_code = 'orange'  # Orange für unter 60%
                    else:
                        korrektur_info = ''
                        sicherheit = 'N/A'
                        seite = ''
                        color_code = 'normal'
                
                # Sortier-Priorität basierend auf Korrektur-Status
                if code in correction_map:
                    sort_priority = 2  # Korrekturen in der Mitte
                else:
                    sort_priority = 3  # Direkte Matches unten
                
                main_results.append({
                    'Code': code,
                    'Status': 'In beiden PDFs',
                    'PDF1': 'Ja',
                    'PDF2': 'Ja',
                    'Sicherheit': sicherheit,
                    'Matching/Korrektur': korrektur_info,
                    'color_code': color_code,
                    'sort_priority': sort_priority,
                    'code_type': 'normal'
                })
            
            
            # 2. Normale Codes nur in PDF1 (rot markiert, aber mit 40% Sicherheit)
            normal_only_pdf1 = corrected_results.get('normal', {}).get('only_in_pdf1', set())
            for code in normal_only_pdf1:
                # Prüfe ob es eine Korrektur für diesen Code gab
                if code in correction_map:
                    corr = correction_map[code]
                    sicherheit = f"{corr['probability']:.0%}"
                    korrektur_info = corr['method']
                    
                    # Aktualisierte Farbkodierung basierend auf Wahrscheinlichkeit
                    if corr['probability'] >= 1.0:
                        color_code = 'dunkel_gruen'  # Dunkel Grün für 100%
                    elif corr['probability'] >= 0.90:
                        color_code = 'hell_gruen'  # Helles Grün für 99-90%
                    elif corr['probability'] >= 0.80:
                        color_code = 'dunkel_gelb'  # Dunkles Gelb für 89-80%
                    elif corr['probability'] >= 0.60:
                        color_code = 'hell_gelb'  # Helles Gelb für 79-60%
                    else:
                        color_code = 'orange'
                else:
                    # Kein Match gefunden: 40% Sicherheit für Fund in Masterliste
                    sicherheit = '40%'
                    korrektur_info = 'Nur in PDF1 gefunden, in Masterliste validiert'
                    color_code = 'rot'  # Normale Codes nur in einem PDF = ROT
                
                main_results.append({
                    'Code': code,
                    'Status': 'Nur in PDF1',
                    'PDF1': 'Ja',
                    'PDF2': 'Nein',
                    'Sicherheit': sicherheit,
                    'Matching/Korrektur': korrektur_info,
                    'color_code': color_code,
                    'sort_priority': 1,  # Kein Match - höchste Priorität (oben)
                    'code_type': 'normal'
                })
            
            # 3. Normale Codes nur in PDF2 (rot markiert, aber mit 40% Sicherheit)
            normal_only_pdf2 = corrected_results.get('normal', {}).get('only_in_pdf2', set())
            for code in normal_only_pdf2:
                # Codes nur in PDF2: 40% Sicherheit für Fund in Masterliste
                sicherheit = '40%'
                korrektur_info = 'Nur in PDF2 gefunden, in Masterliste validiert'
                color_code = 'rot'  # Normale Codes nur in einem PDF = ROT
                
                main_results.append({
                    'Code': code,
                    'Status': 'Nur in PDF2',
                    'PDF1': 'Nein',
                    'PDF2': 'Ja',
                    'Sicherheit': sicherheit,
                    'Matching/Korrektur': korrektur_info,
                    'color_code': color_code,
                    'sort_priority': 1,  # Kein Match - höchste Priorität (oben)
                    'code_type': 'normal'
                })
            
            
            # Sortiere: 0=Steuer-Codes (ganz oben), 1=Normale Abweichungen, 2=Korrekturen, 3=Direkte Matches
            # Innerhalb jeder Kategorie: niedrigste Prozente oben, höchste unten
            def sort_key(x):
                # Extrahiere Prozent-Wert für Sortierung
                if x['Sicherheit'] == 'N/A':
                    percent_value = -1  # Kein Match ganz oben
                else:
                    try:
                        percent_value = float(x['Sicherheit'].replace('%', ''))
                    except:
                        percent_value = 0
                
                return (x['sort_priority'], percent_value, x['Code'])
            
            main_results.sort(key=sort_key)
            
            # Extrahiere Farbinformationen vor DataFrame-Erstellung
            color_mapping = {}
            for idx, result in enumerate(main_results):
                color_mapping[idx] = result['color_code']
            
            # Entferne sort_priority, color_code und code_type für Export
            for result in main_results:
                del result['sort_priority']
                del result['color_code']
                del result['code_type']
            
            df_main = pd.DataFrame(main_results)
            df_main.to_excel(writer, sheet_name='Ergebnisse', index=False)
            
            # Farbkodierung anwenden
            worksheet = writer.sheets['Ergebnisse']
            from openpyxl.styles import PatternFill
            
            # Definiere Farben entsprechend den aktualisierten Anforderungen
            dunkel_gruen_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Dunkles Grün für 100%
            hell_gruen_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")    # Helles Grün für 99-90%
            dunkel_gelb_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # Dunkles Gelb für 89-80%
            hell_gelb_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")     # Helles Gelb für 79-60%
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")        # Orange für unter 60%
            rot_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")          # Rot für nicht gefunden
            grau_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")         # Grau für Steuer-Codes
            
            # Wende Farben auf Zeilen an
            for idx in range(len(main_results)):
                color_code = color_mapping[idx]
                row_idx = idx + 2  # Start bei 2 wegen Header
                
                if color_code == 'dunkel_gruen':
                    fill = dunkel_gruen_fill
                elif color_code == 'hell_gruen':
                    fill = hell_gruen_fill
                elif color_code == 'dunkel_gelb':
                    fill = dunkel_gelb_fill
                elif color_code == 'hell_gelb':
                    fill = hell_gelb_fill
                elif color_code == 'orange':
                    fill = orange_fill
                elif color_code == 'rot':
                    fill = rot_fill
                elif color_code == 'grau':
                    fill = grau_fill
                else:
                    continue
                
                # Färbe die gesamte Zeile
                for col in range(1, len(df_main.columns) + 1):
                    worksheet.cell(row=row_idx, column=col).fill = fill
            
            # Sheet 2: Detaillierte Korrekturen
            if corrections:
                corrections_data = []
                for correction in corrections:
                    if correction['correction_type'] != 'Direkter Match':  # Nur echte Korrekturen (inklusive erweiterte OCR)
                        corrections_data.append({
                            'Originaler_Code': correction['original'],
                            'Korrigierter_Code': correction['corrected'],
                            'Korrektur_Typ': correction['correction_type'],
                            'Seite': correction['page'],
                            'Wahrscheinlichkeit': f"{correction['probability']:.0%}",
                            'Änderungen': correction['method'],
                            'Angewandte_Faktoren': ', '.join(correction.get('factors', []))
                        })
                
                if corrections_data:
                    df_corrections = pd.DataFrame(corrections_data)
                    df_corrections.to_excel(writer, sheet_name='Korrekturen_Details', index=False)
            
            # Sheet 3: Wahrscheinlichkeits-Legende
            if 'legend' in corrected_results or len(corrections) > 0:
                legend_data = []
                # Verwende Legende aus dem ersten Korrektur-Objekt oder Standard-Legende
                legend = corrections[0].get('legend', {}) if corrections else {}
                
                if not legend:
                    # Standard-Legende falls keine verfügbar
                    legend = {
                        "Basis": "50% - Code existiert in Masterliste",
                        "In_PDF2": "+30% - Code wurde auch in PDF2 gefunden",
                        "Direkt_davor": "+10% - Code direkt davor ist identisch in beiden PDFs",
                        "Direkt_danach": "+10% - Code direkt danach ist identisch in beiden PDFs",
                        "Leerzeichen_entfernt": "+20% - Leerzeichen wurden entfernt",
                        "Bekannte_Verwechslung": "+15% - Bekannte OCR-Verwechslung (B↔8, O↔0, etc.)",
                        "Maximum": "100% - Höchstmögliche Sicherheit"
                    }
                
                for factor, description in legend.items():
                    legend_data.append({
                        'Faktor': factor,
                        'Beschreibung': description
                    })
                
                df_legend = pd.DataFrame(legend_data)
                df_legend.to_excel(writer, sheet_name='Wahrscheinlichkeits_Legende', index=False)
            
            # Sheet 4: Zusammenfassung
            actual_corrections = [c for c in corrections if c['correction_type'] != 'Direkter Match']
            advanced_ocr_corrections = [c for c in corrections if c['correction_type'] == 'Erweiterte OCR-Korrektur']
            summary_data = [
                ['Kategorie', 'Original', 'Nach Korrektur', 'Verbesserung'],
                ['In beiden PDFs', len(original_results['in_both']), len(corrected_results['in_both']), 
                 len(corrected_results['in_both']) - len(original_results['in_both'])],
                ['Nur in PDF1', len(original_results['only_in_pdf1']), len(corrected_results['only_in_pdf1']),
                 len(original_results['only_in_pdf1']) - len(corrected_results['only_in_pdf1'])],
                ['Nur in PDF2', len(original_results['only_in_pdf2']), len(corrected_results['only_in_pdf2']),
                 len(original_results['only_in_pdf2']) - len(corrected_results['only_in_pdf2'])],
                ['', '', '', ''],
                ['Durchgeführte Korrekturen', '', len(actual_corrections), ''],
                ['Erweiterte OCR-Korrekturen', '', len(advanced_ocr_corrections), ''],
                ['Leerzeichen-Korrekturen', '', len([c for c in actual_corrections if c['correction_type'] == 'Leerzeichen-Korrektur']), ''],
                ['OCR-Verwechslungen', '', len([c for c in actual_corrections if c['correction_type'] == 'OCR-Verwechslung']), ''],
                ['Verbesserungsrate', '', 
                 f"{len(actual_corrections) / max(len(original_results['only_in_pdf2']), 1) * 100:.1f}%", '']
            ]
            
            df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
            df_summary.to_excel(writer, sheet_name='Zusammenfassung', index=False)
            
            # Formatierung für alle Sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Erweiterter Bericht mit bidirektionaler OCR-Korrektur generiert: {filepath}")
    return filepath